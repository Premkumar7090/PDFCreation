import json
import openpyxl
import pandas as pd
import os
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
import jpype
from openpyxl.utils import get_column_letter, column_index_from_string
from asposecells.api import CellsHelper
from asposecells.api import Workbook, PdfSaveOptions, SaveFormat
import asposecells

from asposecells.api import Workbook, FileFormatType, PdfSaveOptions

# Load the Excel file
wb = openpyxl.load_workbook(r"C:\Users\PRKUMAR\OneDrive - e2open, LLC\PycharmProjects\PDFCitation\BR-CG-CEIS_02262024.xlsx")
df = pd.read_excel(r"C:\Users\PRKUMAR\OneDrive - e2open, LLC\PycharmProjects\PDFCitation\BR-CG-CEIS_02262024.xlsx")







# Get the file name from the path
filename_excel = os.path.splitext(os.path.basename(r"C:\Users\PRKUMAR\OneDrive - e2open, LLC\PycharmProjects\PDFCitation\BR-CG-CEIS_02262024.xlsx"))[0]
# print(filename_excel)
# Read the source workbook into a pandas DataFrame
df_config = pd.read_excel(r"C:\Users\PRKUMAR\OneDrive - e2open, LLC\PycharmProjects\PDFCitation\CITATION_CONFIG.xlsx", sheet_name='Sheet1')

# Check if the search title exists in the specified column
if filename_excel in df_config['RPL_TYPE'].values:
    # Get the row(s) where the search title is found
    matching_rows = df_config[df_config['RPL_TYPE'] == filename_excel]

    # Retrieve the corresponding adjacent row values
  
    adjacent_row_values = matching_rows.iloc[0, 1]
    adjacent_row_values = json.loads(adjacent_row_values)
    print(adjacent_row_values)
else:
    print("Search Title not found in the specified column.")











column_names_dict = {'NOME DO SANCIONADO': 35, 'NOME INFORMADO PELO ÓRGÃO SANCIONADOR':35, 'RAZÃO SOCIAL - CADASTRO RECEITA': 35, 'CATEGORIA DA SANÇÃO': 35, 'ÓRGÃO SANCIONADOR': 35, "DETALHAMENTO":10}
# Dictionary to store column indices
column_indices = {}
# Iterate over the dictionary and check if the column exists in the DataFrame
for column_name, value in adjacent_row_values.items():
   if column_name in df.columns:
       column_index = df.columns.get_loc(column_name)
       column_indices[column_index] = value
print(column_indices)


# Select the sheet
sheet = wb['Sheet1']

column_names = [cell.value for cell in sheet[1]]

# Print the column names
print(column_names)





# Get the worksheet
ws = wb.active

#  Wrap the text in the selected cells and set the border
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True) # Set alignment to top-left
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.row_dimensions[cell.row].height = max(14, int(cell.font.size) + 2)


        # if len(str(cell.value)) > 20:  # Modify the threshold as needed
        #     cell.alignment = cell.alignment.copy(wrapText=True)
        
     

        

# Set the worksheet orientation to landscape
ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE

# Color the header row with yellow background
header_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
for cell in ws[1]:
    cell.fill = header_fill


# Save the Excel file
wb.save('output.xlsx')


workbook = Workbook("output.xlsx")
for sheet in workbook.getWorksheets():
   sheet.autoFitColumns()
   sheet.autoFitRows()
  


# Get the default worksheet
worksheet = workbook.getWorksheets().get(0)









# Set the page orientation to landscape
worksheet.getPageSetup().setOrientation(2)

# Set the paper size to A4

worksheet.getPageSetup().setPaperSize(9)
# Set the maximum rows per page
worksheet.getPageSetup().setFitToPagesTall(1)
worksheet.getPageSetup().setFitToPagesWide(0)
worksheet.setPageBreakPreview(False)


column_indices = column_indices

for index, width in column_indices.items():
    sheet.getCells().setColumnWidth(index, width)

cells = sheet.getCells()
for row in cells.getRows():
    for index in column_indices:
        cell = row.get(index)
        style = cell.getStyle()
        style.setTextWrapped(True)
        cell.setStyle(style)
        
# for sheet in workbook.getWorksheets():
#    sheet.autoFitColumns()
#    sheet.autoFitRows()

# Save as PDF
pdf_options = PdfSaveOptions()
pdf_options.setOnePagePerSheet(True)
# pdf_options.setAllColumnsInOnePagePerSheet(True)
workbook.save("output.pdf", pdf_options)