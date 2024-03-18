import json

import openpyxl
import pandas as pd
import os
import streamlit as st
import tempfile
from openpyxl.styles import Border, Side, Alignment, PatternFill
import jpype
def init_jvm():
    if jpype.isJVMStarted():
        # jpype.shutdownJVM()
        # jpype.startJVM()
        print("JVM is already running")
    else:
        print("JVM Not Started")
        # jpype.shutdownJVM()
        jpype.startJVM()    

init_jvm()


import asposecells

from asposecells.api import PdfSaveOptions
from asposecells.api import Workbook
import datetime

now = datetime.datetime.now()
timestamp = now.timestamp()




# Function to process Excel to PDF
def process_excel_to_pdf(excel_file, config_file):

    
    # Save uploaded Excel file to a temporary location
    with tempfile.NamedTemporaryFile(suffix='.xlsx',delete=False) as temp_excel:
        temp_excel.write(excel_file.read())
        temp_excel_path = temp_excel.name
    column_indices = {}    
    # Load Excel files
    df = pd.read_excel(temp_excel_path)
    print(df.head())
    df_config = pd.read_excel(config_file, sheet_name='Sheet1')
    # Get the filename from the uploaded file
    filename_excel = os.path.splitext(excel_file.name)[0]
    # print(filename_excel)
    
   

    # Check if the search title exists in the specified column
    if filename_excel in df_config['RPL_TYPE'].tolist():
        print("Found in config")
        # Get the row(s) where the search title is found
        matching_rows = df_config[df_config['RPL_TYPE'] == filename_excel]
        # Retrieve the corresponding adjacent row values
        adjacent_row_values = json.loads(matching_rows.iloc[0, 1])
        
        for column_name, value in adjacent_row_values.items():
            if column_name in df.columns:
                column_index = df.columns.get_loc(column_name)
                column_indices[column_index] = value
        print(column_indices)        
    # else:
    #     st.error("Search Title not found in the specified column.")
    #     return None
    # Create a new workbook

    print(temp_excel_path+"-------------------------")
    wb = openpyxl.load_workbook(temp_excel_path)

    # Get the first worksheet
    ws = wb[wb.sheetnames[0]]
    #  Wrap the text in the selected cells and set the border
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True) # Set alignment to top-left
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws.row_dimensions[cell.row].height = max(14, int(cell.font.size) + 2)
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE


    # Color the header row with yellow background
    header_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for cell in ws[1]:
        cell.fill = header_fill


    # Save the Excel file
    wb.save("output.xlsx")  



    workbook = Workbook("output.xlsx")
    for sheet in workbook.getWorksheets():
        sheet.autoFitColumns()
        sheet.autoFitRows()
    
    # Get the default worksheet
    worksheet = workbook.getWorksheets().get(0)
    
    # Set the page orientation to landscape
    worksheet.getPageSetup().setOrientation(2)
    
    # Set the paper size to A4
    
    worksheet.getPageSetup().setPaperSize(9)
    # Set the maximum rows per page
    worksheet.getPageSetup().setFitToPagesTall(1)
    worksheet.getPageSetup().setFitToPagesWide(0)
    worksheet.setPageBreakPreview(False)
    
    
    column_indices = column_indices
    if len(column_indices) != 0:
        print("its inside pdf generation")
        for index, width in column_indices.items():
            sheet.getCells().setColumnWidth(index, width)
        
        cells = sheet.getCells()
        for row in cells.getRows():
            for index in column_indices:
                cell = row.get(index)
                style = cell.getStyle()
                style.setTextWrapped(True)
                cell.setStyle(style)
    else:
        column_indices = [0]
        column_width = 35

        for index in column_indices:
            sheet.getCells().setColumnWidth(index, column_width)
        cells = sheet.getCells()
        for row in cells.getRows():
            for index in column_indices:
                cell = row.get(index)
                style = cell.getStyle()
                style.setTextWrapped(True)
                cell.setStyle(style)       
        
    # Save PDF
    output_pdf_file_path = f"output_{timestamp}.pdf"
    pdf_options = PdfSaveOptions()
    pdf_options.setOnePagePerSheet(True)
    workbook.save(output_pdf_file_path, pdf_options)
    return output_pdf_file_path


# Streamlit UI
st.title("Excel to PDF Converter")
pid = st.text_input('Enter the processor ID: ', '1234567890')

excel_file = st.file_uploader("Upload Excel file", type=["xlsx"])
config_file = st.file_uploader("Upload Config file", type=["xlsx"])
if excel_file and config_file:
    

    


    pdf_file_path = process_excel_to_pdf(excel_file, config_file)
    if pdf_file_path:
        st.success("PDF successfully generated!")
        st.markdown(f"Download [PDF file](/{pdf_file_path})")
        if st.button("Download PDF"):
            st.download_button(label="Download PDF", data=open(pdf_file_path, "rb").read(), file_name=f"output_{timestamp}.pdf", mime="application/pdf")






